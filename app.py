from flask import Flask, render_template, request, jsonify, redirect, send_file
import numpy as np
from numpy.linalg import norm
import pickle
import os
import pymysql
from werkzeug.security import generate_password_hash, check_password_hash
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import re
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

app = Flask(__name__)

# ================================
#  DB 接続設定
# ================================
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",           # 必要に応じて変更
    "database": "care_system",
    "cursorclass": pymysql.cursors.DictCursor,
    "charset": "utf8mb4"
}


def get_connection():
    return pymysql.connect(**DB_CONFIG)


# ================================
#  1. 画面ルーティング
# ================================

@app.route("/")
def index():
    # ここは自由に変えてOK（トップ or テンプレートなど）
    obj = {
        "header_system_name": "認知症初期支援業務管理システム",
        "header_page_name": "テンプレート",
        "footer_sd5": "プロジェクト演習  SD-5"
    }
    return render_template("template.html", d=obj)


@app.route("/top")
def top():
    return render_template("top.html")


@app.route("/list")
def list_page():
    return render_template("list.html")

@app.route("/touroku")
def touroku():
    return render_template("touroku.html")

@app.route("/shousai")
def shousai():
    # ?client_id=1 などのクエリは JS 側で拾って使う
    return render_template("shousai.html")

@app.route("/text")
def text():
    return render_template("text.html")

@app.route("/login", methods=["GET", "POST"])
def login_post():
    if request.method == "POST":
        admin_id = request.form.get("id")
        password = request.form.get("password")

        db = get_connection()
        try:
            cursor = db.cursor()
            cursor.execute("SELECT * FROM admin WHERE admin_id=%s", (admin_id,))
            admin = cursor.fetchone()
        finally:
            db.close()

        if not admin:
            return "ID が存在しません"
        if not check_password_hash(admin['password'], password):
            return "パスワードが違います"

        # ログイン成功 - トップページへリダイレクト
        return redirect("/top")

    # GET の場合はログイン画面を表示
    return render_template("login.html")

@app.route("/register", methods=["POST"])
def register_post():
    # フォームから値を取得
    admin_id = request.form.get("id")
    password = request.form.get("password")
    password2 = request.form.get("password2")

    # 入力チェック
    if not admin_id or not password:
        return "ID またはパスワードが入力されていません"

    if password != password2:
        return "パスワードが一致しません"

    # パスワードをハッシュ化
    hashed = generate_password_hash(password)

    # DB登録
    db = get_connection()
    try:
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO admin (admin_id, password, staff_name) VALUES (%s, %s, %s)",
            (admin_id, hashed, "未設定")
        )
        db.commit()
    except Exception as e:
        db.rollback()
        return f"登録エラー: {e}"
    finally:
        db.close()

    # 登録成功したらログイン画面へ
    return redirect("/login")

# ================================
#  2. テキストマイニング機能（既存）
# ================================
DATA_FILE = "cases.pkl"


def load_cases():
    """過去事例を読み込む（ローカル pickle）"""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "rb") as f:
            return pickle.load(f)
    return []


def save_cases(cases):
    """過去事例を保存"""
    with open(DATA_FILE, "wb") as f:
        pickle.dump(cases, f)


def fake_embedding(text: str):
    """
    ダミーの埋め込み生成
    （同じ文章なら同じベクトルになるよう seed を固定）
    """
    np.random.seed(abs(hash(text)) % (10 ** 7))
    return np.random.rand(128)


def cosine_sim(a, b):
    """コサイン類似度"""
    return float(np.dot(a, b) / (norm(a) * norm(b)))


@app.route("/api/register_case", methods=["POST"])
def register_case():
    """事例登録API（ローカル pickle 保存）"""
    data = request.json
    visit_text = data.get("visit_text", "")
    support_plan = data.get("support_plan", "")

    cases = load_cases()

    new_case = {
        "visit_text": visit_text,
        "support_plan": support_plan,
        "embedding": fake_embedding(visit_text).tolist()
    }

    cases.append(new_case)
    save_cases(cases)

    return jsonify({"status": "ok"})


@app.route("/api/similar_cases", methods=["POST"])
def similar_cases():
    """類似事例検索API"""
    input_text = request.json.get("visit_text", "")
    input_emb = fake_embedding(input_text)

    cases = load_cases()
    results = []

    for case in cases:
        case_emb = np.array(case["embedding"])
        sim = cosine_sim(input_emb, case_emb)
        results.append({
            "similarity": round(sim, 3),
            "visit_text": case["visit_text"],
            "support_plan": case["support_plan"]
        })

    results = sorted(results, key=lambda x: x["similarity"], reverse=True)
    return jsonify(results[:3])


# ================================
#  2.5 TF-IDF ベースのテキストマイニング
# ================================

def extract_keywords(text, top_n=10):
    """テキストからキーワードを抽出（TF-IDF上位語）"""
    if not text or not text.strip():
        return []
    
    words = re.findall(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]+', text)
    if not words:
        return []
    
    word_freq = {}
    for word in words:
        if len(word) >= 2:
            word_freq[word] = word_freq.get(word, 0) + 1
    
    sorted_words = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)
    return [w[0] for w in sorted_words[:top_n]]


def get_visit_records_for_tfidf():
    """DBから訪問記録を取得してTF-IDF用のドキュメントを作成"""
    conn = get_connection()
    records = []
    
    with conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 
                    visit_record_id,
                    client_id,
                    COALESCE(reaction_understanding, '') as reaction,
                    COALESCE(cognitive_function, '') as cognitive,
                    COALESCE(psychiatric_symptoms, '') as psychiatric,
                    COALESCE(physical_condition, '') as physical,
                    COALESCE(living_situation, '') as living,
                    COALESCE(person_wishes, '') as person_wishes,
                    COALESCE(caregiver_wishes, '') as caregiver_wishes,
                    COALESCE(judgment_support, '') as judgment
                FROM visit_record
                WHERE reaction_understanding IS NOT NULL 
                   OR cognitive_function IS NOT NULL
                   OR psychiatric_symptoms IS NOT NULL
            """)
            rows = cur.fetchall()
            
            for row in rows:
                combined_text = " ".join([
                    str(row.get("reaction", "")),
                    str(row.get("cognitive", "")),
                    str(row.get("psychiatric", "")),
                    str(row.get("physical", "")),
                    str(row.get("living", "")),
                    str(row.get("person_wishes", "")),
                    str(row.get("caregiver_wishes", ""))
                ])
                
                if combined_text.strip():
                    records.append({
                        "id": row["visit_record_id"],
                        "client_id": row["client_id"],
                        "text": combined_text,
                        "policy": str(row.get("judgment", "")),
                        "source": "システム入力"
                    })
    
    return records


def get_pdf_case_studies():
    """PDF事例集からのデータを取得"""
    case_file = os.path.join(os.path.dirname(__file__), "data", "case_studies", "cases.json")
    
    if not os.path.exists(case_file):
        return []
    
    try:
        with open(case_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        records = []
        for case in data.get("cases", []):
            records.append({
                "id": case.get("id", ""),
                "client_id": None,
                "text": case.get("text", ""),
                "policy": case.get("policy", ""),
                "source": case.get("source", "PDF事例集"),
                "difficulty_keywords": case.get("difficulty_keywords", []),
                "support_keywords": case.get("support_keywords", [])
            })
        
        return records
    except Exception as e:
        print(f"PDF事例集の読み込みエラー: {e}")
        return []


def get_all_records_for_tfidf():
    """DBの訪問記録とPDF事例集を統合して取得"""
    db_records = get_visit_records_for_tfidf()
    pdf_records = get_pdf_case_studies()
    return db_records + pdf_records


@app.route("/api/search_similar", methods=["POST"])
def search_similar():
    """TF-IDFベースの類似事例検索API（システム入力データ + PDF事例集）"""
    data = request.json or {}
    
    input_parts = [
        data.get("know", ""),
        data.get("ninchi_kinou", ""),
        data.get("symptom", ""),
        data.get("body_con", ""),
        data.get("life_con", ""),
        data.get("honnin_will", ""),
        data.get("kangosha_will", "")
    ]
    input_text = " ".join(input_parts)
    
    if not input_text.strip():
        return jsonify({"results": [], "keywords": []})
    
    keywords = extract_keywords(input_text, top_n=8)
    
    # システム入力データとPDF事例集を統合して取得
    records = get_all_records_for_tfidf()
    
    if not records:
        return jsonify({
            "results": [],
            "keywords": keywords,
            "message": "事例データがありません。訪問記録を登録するか、PDF事例集をインポートしてください。"
        })
    
    corpus = [r["text"] for r in records]
    corpus.append(input_text)
    
    try:
        vectorizer = TfidfVectorizer(
            analyzer='char',
            ngram_range=(2, 4),
            max_features=1000
        )
        tfidf_matrix = vectorizer.fit_transform(corpus)
        
        input_vector = tfidf_matrix[-1]
        doc_vectors = tfidf_matrix[:-1]
        
        similarities = cosine_similarity(input_vector, doc_vectors)[0]
        
        results = []
        for i, sim in enumerate(similarities):
            if sim > 0.01:
                result = {
                    "similarity": round(sim * 100, 1),
                    "text": records[i]["text"][:500],
                    "policy": records[i]["policy"][:500] if records[i]["policy"] else "支援方針未登録",
                    "client_id": records[i].get("client_id"),
                    "source": records[i].get("source", "不明")
                }
                # PDF事例集の場合はキーワード情報も追加
                if records[i].get("difficulty_keywords"):
                    result["difficulty_keywords"] = records[i]["difficulty_keywords"]
                if records[i].get("support_keywords"):
                    result["support_keywords"] = records[i]["support_keywords"]
                results.append(result)
        
        results = sorted(results, key=lambda x: x["similarity"], reverse=True)[:10]
        
    except Exception as e:
        return jsonify({
            "results": [],
            "keywords": keywords,
            "error": str(e)
        })
    
    # 統計情報を追加
    db_count = len([r for r in records if r.get("source") == "システム入力"])
    pdf_count = len([r for r in records if r.get("source") != "システム入力"])
    
    return jsonify({
        "results": results,
        "keywords": keywords,
        "stats": {
            "total_records": len(records),
            "db_records": db_count,
            "pdf_records": pdf_count
        }
    })


# ================================
#  3. アセスメントシート DB API
# ================================

# ------- 共通ユーティリティ -------

def to_int_or_none(value):
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (ValueError, TypeError):
        return None


def to_date_or_none(value):
    """
    <input type="date"> は 'YYYY-MM-DD' または '' が来る想定。
    空なら None を返す（NOT NULL の項目はフロント側で入力を必須にする想定）
    """
    if not value:
        return None
    return value  # そのまま渡す（MySQLが解釈できる形式前提）


def to_datetime_or_none(value):
    """
    datetime-local -> 'YYYY-MM-DDTHH:MM'
    MySQL DATETIME 用に 'YYYY-MM-DD HH:MM:00' に変換
    """
    if not value:
        return None
    v = value.replace("T", " ")
    if len(v) == 16:
        v = v + ":00"
    return v


# ------- client：利用者基本情報 -------

@app.route("/api/save_client", methods=["POST"])
def save_client():
    data = request.json or {}

    client_id = to_int_or_none(data.get("client_id"))

    writer_name = data.get("writer_name", "") or ""
    consultation_date = to_date_or_none(data.get("consultation_date"))
    if consultation_date is None:
        consultation_date = datetime.now().date()
    current_status = data.get("current_status", "") or ""
    client_name = data.get("client_name", "") or ""
    gender = data.get("gender", "") or ""
    birth_date = to_date_or_none(data.get("birth_date"))
    if birth_date is None:
        birth_date = datetime.now().date()
    address = data.get("address", "") or ""
    phone_number = data.get("phone_number", "") or ""
    disability_adl_level = data.get("disability_adl_level", "") or ""
    dementia_adl_level = data.get("dementia_adl_level", "") or ""
    certification_info = data.get("certification_info", "") or ""
    disability_certification = data.get("disability_certification", "") or ""
    living_environment = data.get("living_environment", "") or ""
    economic_status = data.get("economic_status", "") or ""
    visitor_name = data.get("visitor_name", "") or ""
    visitor_contact = data.get("visitor_contact", "") or ""
    relation_to_client = data.get("relation_to_client", "") or ""
    family_composition = data.get("family_composition", "") or ""
    emergency_contact_name = data.get("emergency_contact_name", "") or ""
    emergency_relation = data.get("emergency_relation", "") or ""
    emergency_contact_info = data.get("emergency_contact_info", "") or ""
    life_history = data.get("life_history", "") or ""
    daily_life_pattern = data.get("daily_life_pattern", "") or ""
    time_of_day = data.get("time_of_day", "") or ""
    person_content = data.get("person_content", "") or ""
    caregiver_content = data.get("caregiver_content", "") or ""
    hobbies = data.get("hobbies", "") or ""
    social_connections = data.get("social_connections", "") or ""
    disease_onset_date = to_date_or_none(data.get("disease_onset_date"))
    disease_name = data.get("disease_name", "") or ""
    medical_institution = data.get("medical_institution", "") or ""
    medical_history = data.get("medical_history", "") or ""
    current_condition = data.get("current_condition", "") or ""
    public_services = data.get("public_services", "") or ""
    private_services = data.get("private_services", "") or ""

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            if client_id is not None:
                # 既存かチェック
                cur.execute(
                    "SELECT client_id FROM client WHERE client_id = %s",
                    (client_id,)
                )
                row = cur.fetchone()
            else:
                row = None

            if row:
                # UPDATE
                sql = """
                    UPDATE client
                       SET writer_name=%s,
                           consultation_date=%s,
                           current_status=%s,
                           client_name=%s,
                           gender=%s,
                           birth_date=%s,
                           address=%s,
                           phone_number=%s,
                           disability_adl_level=%s,
                           dementia_adl_level=%s,
                           certification_info=%s,
                           disability_certification=%s,
                           living_environment=%s,
                           economic_status=%s,
                           visitor_name=%s,
                           visitor_contact=%s,
                           relation_to_client=%s,
                           family_composition=%s,
                           emergency_contact_name=%s,
                           emergency_relation=%s,
                           emergency_contact_info=%s,
                           life_history=%s,
                           daily_life_pattern=%s,
                           time_of_day=%s,
                           person_content=%s,
                           caregiver_content=%s,
                           hobbies=%s,
                           social_connections=%s,
                           disease_onset_date=%s,
                           disease_name=%s,
                           medical_institution=%s,
                           medical_history=%s,
                           current_condition=%s,
                           public_services=%s,
                           private_services=%s
                     WHERE client_id=%s
                """
                cur.execute(sql, (
                    writer_name, consultation_date, current_status,
                    client_name, gender, birth_date, address, phone_number,
                    disability_adl_level, dementia_adl_level,
                    certification_info, disability_certification,
                    living_environment, economic_status,
                    visitor_name, visitor_contact, relation_to_client,
                    family_composition, emergency_contact_name,
                    emergency_relation, emergency_contact_info,
                    life_history, daily_life_pattern, time_of_day,
                    person_content, caregiver_content, hobbies,
                    social_connections, disease_onset_date, disease_name,
                    medical_institution, medical_history,
                    current_condition, public_services, private_services,
                    client_id
                ))
            else:
                # INSERT
                sql = """
                    INSERT INTO client (
                        writer_name, consultation_date, current_status,
                        client_name, gender, birth_date, address, phone_number,
                        disability_adl_level, dementia_adl_level,
                        certification_info, disability_certification,
                        living_environment, economic_status,
                        visitor_name, visitor_contact, relation_to_client,
                        family_composition, emergency_contact_name,
                        emergency_relation, emergency_contact_info,
                        life_history, daily_life_pattern, time_of_day,
                        person_content, caregiver_content, hobbies,
                        social_connections, disease_onset_date, disease_name,
                        medical_institution, medical_history,
                        current_condition, public_services, private_services
                    )
                    VALUES (
                        %s,%s,%s,
                        %s,%s,%s,%s,%s,
                        %s,%s,
                        %s,%s,
                        %s,%s,
                        %s,%s,%s,
                        %s,%s,
                        %s,%s,
                        %s,%s,%s,
                        %s,%s,%s,
                        %s,%s,%s,
                        %s,%s,
                        %s,%s,%s
                    )
                """
                cur.execute(sql, (
                    writer_name, consultation_date, current_status,
                    client_name, gender, birth_date, address, phone_number,
                    disability_adl_level, dementia_adl_level,
                    certification_info, disability_certification,
                    living_environment, economic_status,
                    visitor_name, visitor_contact, relation_to_client,
                    family_composition, emergency_contact_name,
                    emergency_relation, emergency_contact_info,
                    life_history, daily_life_pattern, time_of_day,
                    person_content, caregiver_content, hobbies,
                    social_connections, disease_onset_date, disease_name,
                    medical_institution, medical_history,
                    current_condition, public_services, private_services
                ))
                client_id = cur.lastrowid

        conn.commit()

    return jsonify({"status": "saved", "client_id": client_id})


def check_client_exists(client_id):
    """Check if a client exists in the database"""
    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM client WHERE client_id = %s", (client_id,))
            return cur.fetchone() is not None


# ------- visit_record：訪問記録（1クライアント1件を上書き） -------

@app.route("/api/save_visit_record", methods=["POST"])
def save_visit_record():
    data = request.json
    cid = data.get("client_id")

    if not cid:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400

    if not check_client_exists(cid):
        return jsonify({"status": "error", "message": "まず利用者基本情報タブで保存してください"}), 400

    visit_datetime = to_datetime_or_none(data.get("visit_datetime"))
    if visit_datetime is None:
        visit_datetime = datetime.now()
    visitor_name = data.get("visitor_name", "") or ""
    visit_purpose = data.get("visit_purpose", "") or ""
    visit_condition = data.get("visit_condition", "") or ""
    support_decision = data.get("support_decision", "") or ""
    future_plan = data.get("future_plan", "") or ""
    vr_reaction = data.get("vr_reaction", "") or ""
    vr_cognition = data.get("vr_cognition", "") or ""
    vr_dementia_adl = data.get("vr_dementia_adl", "") or ""
    vr_behavior = data.get("vr_behavior", "") or ""
    vr_physical = data.get("vr_physical", "") or ""
    vr_disability_adl = data.get("vr_disability_adl", "") or ""
    vr_living = data.get("vr_living", "") or ""
    vr_dasc = to_int_or_none(data.get("vr_dasc"))
    vr_dbd = to_int_or_none(data.get("vr_dbd"))
    vr_jzbi = to_int_or_none(data.get("vr_jzbi"))
    vr_person_intent = data.get("vr_person_intent", "") or ""
    vr_family_intent = data.get("vr_family_intent", "") or ""
    vr_other = data.get("vr_other", "") or ""

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:

            cur.execute("SELECT visit_record_id FROM visit_record WHERE client_id=%s", (cid,))
            row = cur.fetchone()

            if row:
                sql = """
                    UPDATE visit_record SET
                        visit_datetime=%s,
                        visitor_name=%s,
                        visit_purpose=%s,
                        visit_condition=%s,
                        support_decision=%s,
                        future_plan=%s,
                        vr_reaction=%s,
                        vr_cognition=%s,
                        vr_dementia_adl=%s,
                        vr_behavior=%s,
                        vr_physical=%s,
                        vr_disability_adl=%s,
                        vr_living=%s,
                        vr_dasc=%s,
                        vr_dbd=%s,
                        vr_jzbi=%s,
                        vr_person_intent=%s,
                        vr_family_intent=%s,
                        vr_other=%s
                    WHERE client_id=%s
                """
                cur.execute(sql, (
                    visit_datetime,
                    visitor_name,
                    visit_purpose,
                    visit_condition,
                    support_decision,
                    future_plan,
                    vr_reaction,
                    vr_cognition,
                    vr_dementia_adl,
                    vr_behavior,
                    vr_physical,
                    vr_disability_adl,
                    vr_living,
                    vr_dasc,
                    vr_dbd,
                    vr_jzbi,
                    vr_person_intent,
                    vr_family_intent,
                    vr_other,
                    cid
                ))
            else:
                sql = """
                    INSERT INTO visit_record (
                        client_id, visit_datetime, visitor_name,
                        visit_purpose, visit_condition,
                        support_decision, future_plan,
                        vr_reaction, vr_cognition, vr_dementia_adl,
                        vr_behavior, vr_physical, vr_disability_adl,
                        vr_living, vr_dasc, vr_dbd, vr_jzbi,
                        vr_person_intent, vr_family_intent, vr_other
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cur.execute(sql, (
                    cid,
                    visit_datetime,
                    visitor_name,
                    visit_purpose,
                    visit_condition,
                    support_decision,
                    future_plan,
                    vr_reaction,
                    vr_cognition,
                    vr_dementia_adl,
                    vr_behavior,
                    vr_physical,
                    vr_disability_adl,
                    vr_living,
                    vr_dasc,
                    vr_dbd,
                    vr_jzbi,
                    vr_person_intent,
                    vr_family_intent,
                    vr_other
                ))

        conn.commit()

    return jsonify({"status": "saved"})


# ------- physical_status：身体状況チェック（1クライアント1件上書き） -------

@app.route("/api/save_physical_status", methods=["POST"])
def save_physical_status():
    data = request.json
    cid = data.get("client_id")

    if not cid:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400

    if not check_client_exists(cid):
        return jsonify({"status": "error", "message": "まず利用者基本情報タブで保存してください"}), 400

    ps_mobility = data.get("ps_mobility", "") or ""
    ps_walking = data.get("ps_walking", "") or ""
    ps_transport = data.get("ps_transport", "") or ""
    ps_communication = data.get("ps_communication", "") or ""
    ps_decision = data.get("ps_decision", "") or ""
    ps_senses = data.get("ps_senses", "") or ""
    ps_hygiene = data.get("ps_hygiene", "") or ""
    ps_cleanliness = data.get("ps_cleanliness", "") or ""
    ps_nutrition = data.get("ps_nutrition", "") or ""
    ps_eating_behavior = data.get("ps_eating_behavior", "") or ""
    ps_swallowing = data.get("ps_swallowing", "") or ""
    ps_meal_refusal = data.get("ps_meal_refusal", "") or ""
    ps_water = data.get("ps_water", "") or ""
    ps_habits = data.get("ps_habits", "") or ""
    ps_excretion = data.get("ps_excretion", "") or ""
    ps_constipation = data.get("ps_constipation", "") or ""
    ps_sleep = data.get("ps_sleep", "") or ""
    ps_daily_rhythm = data.get("ps_daily_rhythm", "") or ""
    ps_daytime_sleep = data.get("ps_daytime_sleep", "") or ""
    ps_night_behavior = data.get("ps_night_behavior", "") or ""
    ps_house_env = data.get("ps_house_env", "") or ""
    ps_money = data.get("ps_money", "") or ""
    ps_family_care = data.get("ps_family_care", "") or ""
    ps_abuse = data.get("ps_abuse", "") or ""
    ps_watch = data.get("ps_watch", "") or ""
    ps_sos = data.get("ps_sos", "") or ""

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:

            cur.execute("SELECT physical_status_id FROM physical_status WHERE client_id=%s", (cid,))
            row = cur.fetchone()

            if row:
                sql = """
                    UPDATE physical_status SET
                        ps_mobility=%s, ps_walking=%s, ps_transport=%s,
                        ps_communication=%s, ps_decision=%s, ps_senses=%s,
                        ps_hygiene=%s, ps_cleanliness=%s,
                        ps_nutrition=%s, ps_eating_behavior=%s, ps_swallowing=%s, ps_meal_refusal=%s,
                        ps_water=%s, ps_habits=%s,
                        ps_excretion=%s, ps_constipation=%s,
                        ps_sleep=%s, ps_daily_rhythm=%s, ps_daytime_sleep=%s, ps_night_behavior=%s,
                        ps_house_env=%s, ps_money=%s, ps_family_care=%s, ps_abuse=%s,
                        ps_watch=%s, ps_sos=%s
                    WHERE client_id=%s
                """
                cur.execute(sql, (
                    ps_mobility, ps_walking, ps_transport,
                    ps_communication, ps_decision, ps_senses,
                    ps_hygiene, ps_cleanliness,
                    ps_nutrition, ps_eating_behavior, ps_swallowing, ps_meal_refusal,
                    ps_water, ps_habits,
                    ps_excretion, ps_constipation,
                    ps_sleep, ps_daily_rhythm, ps_daytime_sleep, ps_night_behavior,
                    ps_house_env, ps_money, ps_family_care, ps_abuse,
                    ps_watch, ps_sos,
                    cid
                ))
            else:
                sql = """
                    INSERT INTO physical_status (
                        client_id,
                        ps_mobility, ps_walking, ps_transport,
                        ps_communication, ps_decision, ps_senses,
                        ps_hygiene, ps_cleanliness,
                        ps_nutrition, ps_eating_behavior, ps_swallowing, ps_meal_refusal,
                        ps_water, ps_habits,
                        ps_excretion, ps_constipation,
                        ps_sleep, ps_daily_rhythm, ps_daytime_sleep, ps_night_behavior,
                        ps_house_env, ps_money, ps_family_care, ps_abuse,
                        ps_watch, ps_sos
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cur.execute(sql, (
                    cid,
                    ps_mobility, ps_walking, ps_transport,
                    ps_communication, ps_decision, ps_senses,
                    ps_hygiene, ps_cleanliness,
                    ps_nutrition, ps_eating_behavior, ps_swallowing, ps_meal_refusal,
                    ps_water, ps_habits,
                    ps_excretion, ps_constipation,
                    ps_sleep, ps_daily_rhythm, ps_daytime_sleep, ps_night_behavior,
                    ps_house_env, ps_money, ps_family_care, ps_abuse,
                    ps_watch, ps_sos
                ))

        conn.commit()

    return jsonify({"status": "saved"})


# ------- dasc21：DASC-21（1クライアント1件を上書き） -------

@app.route("/api/save_dasc21", methods=["POST"])
def save_dasc21():
    data = request.json or {}

    client_id = to_int_or_none(data.get("client_id"))
    if client_id is None:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400

    if not check_client_exists(client_id):
        return jsonify({"status": "error", "message": "まず利用者基本情報タブで保存してください"}), 400

    informant_name = data.get("informant_name", "") or ""
    evaluator_name = data.get("evaluator_name", "") or ""
    assessment_item = data.get("assessment_item", "") or ""
    remarks = data.get("remarks", "") or ""
    total_score = to_int_or_none(data.get("total_score")) or 0

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT dasc_id FROM dasc21 WHERE client_id=%s ORDER BY dasc_id DESC LIMIT 1",
                (client_id,)
            )
            row = cur.fetchone()

            if row:
                dasc_id = row["dasc_id"]
                sql = """
                    UPDATE dasc21 SET
                        informant_name=%s,
                        evaluator_name=%s,
                        assessment_item=%s,
                        remarks=%s,
                        total_score=%s
                    WHERE dasc_id=%s
                """
                cur.execute(sql, (
                    informant_name, evaluator_name,
                    assessment_item, remarks, total_score,
                    dasc_id
                ))
            else:
                sql = """
                    INSERT INTO dasc21 (
                        client_id, informant_name, evaluator_name,
                        assessment_item, remarks, total_score
                    )
                    VALUES (%s,%s,%s,%s,%s,%s)
                """
                cur.execute(sql, (
                    client_id, informant_name, evaluator_name,
                    assessment_item, remarks, total_score
                ))
                dasc_id = cur.lastrowid
        conn.commit()

    return jsonify({"status": "saved", "dasc_id": dasc_id})


# ------- dbd13：DBD-13（1クライアント1件を上書き） -------

@app.route("/api/save_dbd13", methods=["POST"])
def save_dbd13():
    data = request.json or {}

    client_id = to_int_or_none(data.get("client_id"))
    if client_id is None:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400

    if not check_client_exists(client_id):
        return jsonify({"status": "error", "message": "まず利用者基本情報タブで保存してください"}), 400

    respondent_name = data.get("respondent_name", "") or ""
    evaluator_name = data.get("evaluator_name", "") or ""
    entry_date = to_date_or_none(data.get("entry_date"))
    if entry_date is None:
        entry_date = datetime.now().date()
    assessment_item = data.get("assessment_item", "") or ""
    remarks = data.get("remarks", "") or ""
    subtotal_score = to_int_or_none(data.get("subtotal_score")) or 0
    total_score = to_int_or_none(data.get("total_score")) or 0

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT dbd_id FROM dbd13 WHERE client_id=%s ORDER BY dbd_id DESC LIMIT 1",
                (client_id,)
            )
            row = cur.fetchone()

            if row:
                dbd_id = row["dbd_id"]
                sql = """
                    UPDATE dbd13 SET
                        respondent_name=%s,
                        evaluator_name=%s,
                        entry_date=%s,
                        assessment_item=%s,
                        remarks=%s,
                        subtotal_score=%s,
                        total_score=%s
                    WHERE dbd_id=%s
                """
                cur.execute(sql, (
                    respondent_name, evaluator_name, entry_date,
                    assessment_item, remarks, subtotal_score, total_score,
                    dbd_id
                ))
            else:
                sql = """
                    INSERT INTO dbd13 (
                        client_id, respondent_name, evaluator_name, entry_date,
                        assessment_item, remarks, subtotal_score, total_score
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """
                cur.execute(sql, (
                    client_id, respondent_name, evaluator_name, entry_date,
                    assessment_item, remarks, subtotal_score, total_score
                ))
                dbd_id = cur.lastrowid
        conn.commit()

    return jsonify({"status": "saved", "dbd_id": dbd_id})


# ------- 対象者一覧取得API -------

@app.route("/api/get_clients")
def get_clients():
    """対象者一覧を取得"""
    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT client_id, client_name, gender, consultation_date, current_status
                FROM client
                ORDER BY client_id DESC
            """)
            clients = cur.fetchall()

    return jsonify({"status": "ok", "clients": clients})


# ------- shousai 画面用：一括取得API -------

@app.route("/api/get_all_data")
def get_all_data():
    client_id = to_int_or_none(request.args.get("client_id"))
    if client_id is None:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            # client
            cur.execute("SELECT * FROM client WHERE client_id=%s", (client_id,))
            client = cur.fetchone()

            # visit_record（最新1件）
            cur.execute(
                "SELECT * FROM visit_record WHERE client_id=%s ORDER BY visit_record_id DESC LIMIT 1",
                (client_id,)
            )
            visit_record = cur.fetchone()

            # physical_status（1件）
            cur.execute(
                "SELECT * FROM physical_status WHERE client_id=%s ORDER BY physical_status_id DESC LIMIT 1",
                (client_id,)
            )
            physical_status = cur.fetchone()

            # dasc21（最新1件）
            cur.execute(
                "SELECT * FROM dasc21 WHERE client_id=%s ORDER BY dasc_id DESC LIMIT 1",
                (client_id,)
            )
            dasc21_row = cur.fetchone()

            # dbd13（最新1件）
            cur.execute(
                "SELECT * FROM dbd13 WHERE client_id=%s ORDER BY dbd_id DESC LIMIT 1",
                (client_id,)
            )
            dbd13_row = cur.fetchone()

    return jsonify({
        "status": "ok",
        "client": client,
        "visit_record": visit_record,
        "physical_status": physical_status,
        "dasc21": dasc21_row,
        "dbd13": dbd13_row
    })

# ================================
#  4. 共有フォルダ（ファイル管理）
# ================================

from flask import send_from_directory
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = "uploads"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# フォルダなければ作成
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# -------------------------
# ⑥ ファイルアップロード
# -------------------------
@app.route("/api/upload_file", methods=["POST"])
def upload_file():
    client_id = request.form.get("client_id")
    file = request.files.get("file")

    if not client_id:
        return jsonify({"status": "error", "message": "client_id がありません"}), 400
    if not file:
        return jsonify({"status": "error", "message": "ファイルがありません"}), 400

    # 保存先：/uploads/<client_id>/
    client_folder = os.path.join(app.config["UPLOAD_FOLDER"], str(client_id))
    os.makedirs(client_folder, exist_ok=True)

    filename = secure_filename(file.filename)
    save_path = os.path.join(client_folder, filename)
    file.save(save_path)

    # DB 登録
    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            sql = """
                INSERT INTO shared_folder (client_id, file_path)
                VALUES (%s, %s)
            """
            cur.execute(sql, (client_id, f"{client_id}/{filename}"))
        conn.commit()

    return jsonify({"status": "saved"})


# -------------------------
# ⑦ ファイル一覧取得
# -------------------------
@app.route("/api/files", methods=["GET"])
def get_files():
    client_id = request.args.get("client_id")

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT folder_id, file_path, uploaded_at FROM shared_folder WHERE client_id=%s",
                (client_id,)
            )
            files = cur.fetchall()

    return jsonify({"status": "ok", "files": files})


# -------------------------
# ⑧ ファイル削除
# -------------------------
@app.route("/api/delete_file", methods=["POST"])
def delete_file():
    folder_id = request.json.get("folder_id")

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT file_path FROM shared_folder WHERE folder_id=%s", (folder_id,))
            row = cur.fetchone()

            if not row:
                return jsonify({"status": "error", "message": "ファイルが存在しません"}), 404

            file_path = row["file_path"]

            # DB から削除
            cur.execute("DELETE FROM shared_folder WHERE folder_id=%s", (folder_id,))
        conn.commit()

    # 実ファイルも削除
    actual = os.path.join(app.config["UPLOAD_FOLDER"], file_path)
    if os.path.exists(actual):
        os.remove(actual)

    return jsonify({"status": "deleted"})


# -------------------------
# ⑨ ファイルダウンロード
# -------------------------
@app.route("/uploads/<path:filepath>")
def uploaded_file(filepath):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filepath, as_attachment=True)

# ================================
#  5. Excel出力API
# ================================

def save_excel_to_client_folder(wb, client_id, sheet_type):
    """ExcelファイルをクライアントフォルダにSaveし、DBに登録"""
    # uploadsフォルダに保存（フォルダUIと同じ場所）
    client_folder = os.path.join(app.config["UPLOAD_FOLDER"], str(client_id))
    os.makedirs(client_folder, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{sheet_type}_{client_id}_{timestamp}.xlsx"
    filepath = os.path.join(client_folder, filename)
    wb.save(filepath)
    
    # DBのshared_folderテーブルに登録（フォルダUIに表示されるように）
    relative_path = f"{client_id}/{filename}"
    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO shared_folder (client_id, file_path) VALUES (%s, %s)",
                (client_id, relative_path)
            )
        conn.commit()
    
    return filepath

def create_excel_styles():
    """共通のExcelスタイルを作成"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8F4FC", end_color="E8F4FC", fill_type="solid")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    # セクションヘッダー用スタイル（shousai.htmlのassessment-section-headerに対応）
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'title_font': title_font,
        'title_fill': title_fill,
        'section_font': section_font,
        'section_fill': section_fill,
        'thin_border': thin_border,
        'center_align': center_align,
        'left_align': left_align
    }


def add_section_header(ws, row, title, styles, col_span=2):
    """セクションヘッダーを追加するヘルパー関数"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles['section_font']
    cell.fill = styles['section_fill']
    cell.alignment = styles['center_align']
    cell.border = styles['thin_border']
    ws.row_dimensions[row].height = 25
    return row + 1

def add_field_row(ws, row, label, value, styles):
    """フィールド行を追加するヘルパー関数"""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = styles['header_font']
    label_cell.fill = styles['header_fill']
    label_cell.border = styles['thin_border']
    label_cell.alignment = styles['left_align']
    
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.border = styles['thin_border']
    value_cell.alignment = styles['left_align']
    return row + 1

@app.route("/api/export_client", methods=["GET"])
def export_client():
    """利用者基本情報をExcelで出力"""
    client_id = request.args.get("client_id")
    if not client_id:
        return jsonify({"status": "error", "message": "client_id が必要です"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM client WHERE client_id = %s", (client_id,))
            data = cur.fetchone()

    if not data:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "利用者基本情報"
    styles = create_excel_styles()

    # タイトル行
    ws.merge_cells('A1:B1')
    ws['A1'] = "利用者基本情報"
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws['A1'].alignment = styles['center_align']
    ws.row_dimensions[1].height = 30

    row = 3
    
    # 基本情報セクション
    row = add_section_header(ws, row, "基本情報", styles)
    row = add_field_row(ws, row, "作成担当者", data.get("writer_name", ""), styles)
    row = add_field_row(ws, row, "相談日", str(data.get("consultation_date", "")) if data.get("consultation_date") else "", styles)
    row = add_field_row(ws, row, "本人氏名", data.get("client_name", ""), styles)
    row = add_field_row(ws, row, "性別", data.get("gender", ""), styles)
    row = add_field_row(ws, row, "生年月日", str(data.get("birth_date", "")) if data.get("birth_date") else "", styles)
    row = add_field_row(ws, row, "電話番号", data.get("phone_number", ""), styles)
    row = add_field_row(ws, row, "住所", data.get("address", ""), styles)
    row = add_field_row(ws, row, "本人の現況", data.get("current_status", ""), styles)
    
    row += 1  # 空行
    
    # 自立度・認定情報セクション
    row = add_section_header(ws, row, "自立度・認定情報", styles)
    row = add_field_row(ws, row, "障害高齢者の日常生活自立度", data.get("disability_adl_level", ""), styles)
    row = add_field_row(ws, row, "認知症高齢者の日常生活自立度", data.get("dementia_adl_level", ""), styles)
    row = add_field_row(ws, row, "認定・総合事業情報", data.get("certification_info", ""), styles)
    row = add_field_row(ws, row, "障害等認定", data.get("disability_certification", ""), styles)
    
    row += 1  # 空行
    
    # 生活環境セクション
    row = add_section_header(ws, row, "生活環境", styles)
    row = add_field_row(ws, row, "住居環境", data.get("living_environment", ""), styles)
    row = add_field_row(ws, row, "経済状況", data.get("economic_status", ""), styles)
    
    row += 1  # 空行
    
    # 相談者・家族情報セクション
    row = add_section_header(ws, row, "相談者・家族情報", styles)
    row = add_field_row(ws, row, "来所者・相談者氏名", data.get("visitor_name", ""), styles)
    row = add_field_row(ws, row, "本人との続柄", data.get("relation_to_client", ""), styles)
    row = add_field_row(ws, row, "来所者連絡先", data.get("visitor_contact", ""), styles)
    row = add_field_row(ws, row, "家族構成", data.get("family_composition", ""), styles)
    row = add_field_row(ws, row, "緊急連絡先氏名", data.get("emergency_contact_name", ""), styles)
    row = add_field_row(ws, row, "緊急連絡先続柄", data.get("emergency_relation", ""), styles)
    row = add_field_row(ws, row, "緊急連絡先", data.get("emergency_contact_info", ""), styles)
    
    row += 1  # 空行
    
    # 生活歴・日常生活セクション
    row = add_section_header(ws, row, "生活歴・日常生活", styles)
    row = add_field_row(ws, row, "生活歴", data.get("life_history", ""), styles)
    row = add_field_row(ws, row, "日常生活パターン", data.get("daily_life_pattern", ""), styles)
    row = add_field_row(ws, row, "時間帯", data.get("time_of_day", ""), styles)
    row = add_field_row(ws, row, "本人の内容", data.get("person_content", ""), styles)
    row = add_field_row(ws, row, "介護者の内容", data.get("caregiver_content", ""), styles)
    
    row += 1  # 空行
    
    # 趣味・社会参加セクション
    row = add_section_header(ws, row, "趣味・社会参加", styles)
    row = add_field_row(ws, row, "趣味・嗜好", data.get("hobbies", ""), styles)
    row = add_field_row(ws, row, "社会的つながり", data.get("social_connections", ""), styles)
    
    row += 1  # 空行
    
    # 医療情報セクション
    row = add_section_header(ws, row, "医療情報", styles)
    row = add_field_row(ws, row, "発症日", str(data.get("disease_onset_date", "")) if data.get("disease_onset_date") else "", styles)
    row = add_field_row(ws, row, "疾患名", data.get("disease_name", ""), styles)
    row = add_field_row(ws, row, "主治医・医療機関", data.get("medical_institution", ""), styles)
    row = add_field_row(ws, row, "既往歴", data.get("medical_history", ""), styles)
    row = add_field_row(ws, row, "現在の状態・経過", data.get("current_condition", ""), styles)
    
    row += 1  # 空行
    
    # サービス利用状況セクション
    row = add_section_header(ws, row, "サービス利用状況", styles)
    row = add_field_row(ws, row, "現在利用中の公的サービス", data.get("public_services", ""), styles)
    row = add_field_row(ws, row, "現在利用中の非公的サービス", data.get("private_services", ""), styles)

    # 列幅を広げて文字がはみ出さないようにする
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 60

    save_excel_to_client_folder(wb, client_id, "client")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"client_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route("/api/export_visit", methods=["GET"])
def export_visit():
    """訪問記録表をExcelで出力"""
    client_id = request.args.get("client_id")
    if not client_id:
        return jsonify({"status": "error", "message": "client_id が必要です"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM visit_record WHERE client_id = %s ORDER BY visit_datetime DESC LIMIT 1", (client_id,))
            data = cur.fetchone()

    if not data:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "訪問記録表"
    styles = create_excel_styles()

    # タイトル行
    ws.merge_cells('A1:B1')
    ws['A1'] = "訪問記録表"
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws['A1'].alignment = styles['center_align']
    ws.row_dimensions[1].height = 30

    row = 3
    
    # 訪問基本情報セクション
    row = add_section_header(ws, row, "訪問基本情報", styles)
    row = add_field_row(ws, row, "訪問日時", str(data.get("visit_datetime", "")) if data.get("visit_datetime") else "", styles)
    row = add_field_row(ws, row, "訪問者氏名", data.get("visitor_name", ""), styles)
    row = add_field_row(ws, row, "訪問目的", data.get("visit_purpose", ""), styles)
    row = add_field_row(ws, row, "訪問時の状態", data.get("visit_condition", ""), styles)
    
    row += 1  # 空行
    
    # 評価内容セクション
    row = add_section_header(ws, row, "評価内容", styles)
    row = add_field_row(ws, row, "訪問に対する本人の反応・理解", data.get("vr_reaction", ""), styles)
    row = add_field_row(ws, row, "認知機能", data.get("vr_cognition", ""), styles)
    row = add_field_row(ws, row, "認知症日常生活自立度", data.get("vr_dementia_adl", ""), styles)
    row = add_field_row(ws, row, "精神症状・行動症状", data.get("vr_behavior", ""), styles)
    row = add_field_row(ws, row, "身体状況", data.get("vr_physical", ""), styles)
    row = add_field_row(ws, row, "障害高齢者の日常生活自立度", data.get("vr_disability_adl", ""), styles)
    row = add_field_row(ws, row, "生活状況", data.get("vr_living", ""), styles)
    
    row += 1  # 空行
    
    # 評価点数セクション
    row = add_section_header(ws, row, "評価点数", styles)
    row = add_field_row(ws, row, "DASC-21 点数", str(data.get("vr_dasc", "")) if data.get("vr_dasc") else "", styles)
    row = add_field_row(ws, row, "DBD13 点数", str(data.get("vr_dbd", "")) if data.get("vr_dbd") else "", styles)
    row = add_field_row(ws, row, "J-ZBI8 点数", str(data.get("vr_jzbi", "")) if data.get("vr_jzbi") else "", styles)
    
    row += 1  # 空行
    
    # 意向・希望セクション
    row = add_section_header(ws, row, "意向・希望", styles)
    row = add_field_row(ws, row, "本人の意向・希望", data.get("vr_person_intent", ""), styles)
    row = add_field_row(ws, row, "介護者の意向・希望", data.get("vr_family_intent", ""), styles)
    row = add_field_row(ws, row, "その他", data.get("vr_other", ""), styles)
    
    row += 1  # 空行
    
    # 支援計画セクション
    row = add_section_header(ws, row, "支援計画", styles)
    row = add_field_row(ws, row, "判断・支援内容", data.get("support_decision", ""), styles)
    row = add_field_row(ws, row, "今後の方針・支援計画", data.get("future_plan", ""), styles)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 60

    save_excel_to_client_folder(wb, client_id, "visit")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"visit_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route("/api/export_physical", methods=["GET"])
def export_physical():
    """身体状況チェック表をExcelで出力"""
    client_id = request.args.get("client_id")
    if not client_id:
        return jsonify({"status": "error", "message": "client_id が必要です"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM physical_status WHERE client_id = %s ORDER BY physical_status_id DESC LIMIT 1", (client_id,))
            data = cur.fetchone()

    if not data:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "身体状況チェック表"
    styles = create_excel_styles()

    # タイトル行
    ws.merge_cells('A1:B1')
    ws['A1'] = "身体状況チェック表"
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws['A1'].alignment = styles['center_align']
    ws.row_dimensions[1].height = 30

    row = 3
    
    # 運動機能セクション
    row = add_section_header(ws, row, "運動機能", styles)
    row = add_field_row(ws, row, "立ち上がり・運動機能", data.get("ps_mobility", "") or "", styles)
    row = add_field_row(ws, row, "歩行状況", data.get("ps_walking", "") or "", styles)
    row = add_field_row(ws, row, "移動範囲", data.get("ps_transport", "") or "", styles)
    
    row += 1  # 空行
    
    # 認知・コミュニケーションセクション
    row = add_section_header(ws, row, "認知・コミュニケーション", styles)
    row = add_field_row(ws, row, "意思疎通", data.get("ps_communication", "") or "", styles)
    row = add_field_row(ws, row, "意思決定能力", data.get("ps_decision", "") or "", styles)
    row = add_field_row(ws, row, "視力・聴力", data.get("ps_senses", "") or "", styles)
    
    row += 1  # 空行
    
    # 清潔・衛生セクション
    row = add_section_header(ws, row, "清潔・衛生", styles)
    row = add_field_row(ws, row, "入浴と清潔状態", data.get("ps_hygiene", "") or "", styles)
    row = add_field_row(ws, row, "衣類・家屋の清潔さ", data.get("ps_cleanliness", "") or "", styles)
    
    row += 1  # 空行
    
    # 食事・栄養セクション
    row = add_section_header(ws, row, "食事・栄養", styles)
    row = add_field_row(ws, row, "栄養状態", data.get("ps_nutrition", "") or "", styles)
    row = add_field_row(ws, row, "過食・異食", data.get("ps_eating_behavior", "") or "", styles)
    row = add_field_row(ws, row, "嚥下能力", data.get("ps_swallowing", "") or "", styles)
    row = add_field_row(ws, row, "食事拒否・時間", data.get("ps_meal_refusal", "") or "", styles)
    row = add_field_row(ws, row, "水分摂取状況", data.get("ps_water", "") or "", styles)
    row = add_field_row(ws, row, "飲酒と喫煙", data.get("ps_habits", "") or "", styles)
    
    row += 1  # 空行
    
    # 排泄セクション
    row = add_section_header(ws, row, "排泄", styles)
    row = add_field_row(ws, row, "排泄状況", data.get("ps_excretion", "") or "", styles)
    row = add_field_row(ws, row, "便秘（下剤）", data.get("ps_constipation", "") or "", styles)
    
    row += 1  # 空行
    
    # 睡眠・生活リズムセクション
    row = add_section_header(ws, row, "睡眠・生活リズム", styles)
    row = add_field_row(ws, row, "睡眠状況", data.get("ps_sleep", "") or "", styles)
    row = add_field_row(ws, row, "生活リズム", data.get("ps_daily_rhythm", "") or "", styles)
    row = add_field_row(ws, row, "日中の睡眠", data.get("ps_daytime_sleep", "") or "", styles)
    row = add_field_row(ws, row, "夜間の行動", data.get("ps_night_behavior", "") or "", styles)
    
    row += 1  # 空行
    
    # 生活環境・支援状況セクション
    row = add_section_header(ws, row, "生活環境・支援状況", styles)
    row = add_field_row(ws, row, "居住環境の問題", data.get("ps_house_env", "") or "", styles)
    row = add_field_row(ws, row, "金銭管理", data.get("ps_money", "") or "", styles)
    row = add_field_row(ws, row, "家族の介護力", data.get("ps_family_care", "") or "", styles)
    row = add_field_row(ws, row, "虐待可能性", data.get("ps_abuse", "") or "", styles)
    row = add_field_row(ws, row, "見守り状況", data.get("ps_watch", "") or "", styles)
    row = add_field_row(ws, row, "SOS発信可否", data.get("ps_sos", "") or "", styles)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 60

    save_excel_to_client_folder(wb, client_id, "physical")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"physical_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route("/api/export_dasc21", methods=["GET"])
def export_dasc21():
    """DASC-21をExcelで出力"""
    client_id = request.args.get("client_id")
    if not client_id:
        return jsonify({"status": "error", "message": "client_id が必要です"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM dasc21 WHERE client_id = %s ORDER BY dasc_id DESC LIMIT 1", (client_id,))
            data = cur.fetchone()

    if not data:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    import json
    assessment_answers = {}
    try:
        assessment_item = data.get("assessment_item", "")
        if assessment_item:
            assessment_answers = json.loads(assessment_item)
    except (json.JSONDecodeError, TypeError):
        assessment_answers = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "DASC-21"
    styles = create_excel_styles()

    ws.merge_cells('A1:G1')
    ws['A1'] = "DASC-21 地域包括ケアシステムにおける認知症アセスメントシート"
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws['A1'].alignment = styles['center_align']
    ws.row_dimensions[1].height = 30

    ws['A3'] = "情報提供者氏名"
    ws['B3'] = data.get("informant_name", "")
    ws['C3'] = "評価者氏名"
    ws['D3'] = data.get("evaluator_name", "")
    ws['E3'] = "評価日"
    ws['F3'] = assessment_answers.get("evaluation_date", "")
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].border = styles['thin_border']

    ws['A5'] = "評価基準: 1=問題なくできる 2=だいたいできる 3=あまりできない 4=できない"
    ws['A5'].font = Font(italic=True, size=10)

    dasc_items = [
        ("A. 記憶", [
            ("1", "財布や鍵など、物を置いた場所がわからなくなることがありますか"),
            ("2", "5分前に聞いた話を思い出せないことがありますか"),
            ("3", "自分の生年月日がわからなくなることがありますか"),
        ]),
        ("B. 見当識", [
            ("4", "今日が何月何日かわからないときがありますか"),
            ("5", "自分のいる場所がどこだかわからなくなることがありますか"),
            ("6", "道に迷って家に帰ってこられなくなることがありますか"),
        ]),
        ("C. 問題解決・判断力", [
            ("7", "電気やガスや水道が止まってしまったときに、自分で適切に対処できますか"),
            ("8", "一日の計画を自分で立てることができますか"),
            ("9", "季節や状況に合った服を自分で選ぶことができますか"),
        ]),
        ("D. 家庭外のIADL", [
            ("10", "バスや電車、自家用車などを使って一人で外出できますか"),
            ("11", "貯金の出し入れや、家賃や公共料金の支払いは一人でできますか"),
        ]),
        ("E. 家庭内のIADL", [
            ("12", "薬を決まった時間に決まった分量のむことはできますか"),
            ("13", "電話をかけることができますか"),
            ("14", "自分で食事の準備はできますか"),
            ("15", "自分で、掃除機やほうきを使って掃除ができますか"),
            ("16", "自分で洗濯ができますか"),
        ]),
        ("F. 身体的ADL", [
            ("17", "自分で適切な量の食事をとることはできますか"),
            ("18", "入浴は一人でできますか"),
            ("19", "トイレは一人でできますか"),
            ("20", "身だしなみを整えることは一人でできますか"),
            ("21", "一人で外出できますか"),
        ]),
    ]

    row = 7
    headers = ["No.", "評価項目", "1", "2", "3", "4", "評価"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
    row += 1

    for category, items in dasc_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=category)
        cell.font = Font(bold=True, color="2C5282")
        cell.fill = PatternFill(start_color="E8F4FC", end_color="E8F4FC", fill_type="solid")
        cell.border = styles['thin_border']
        row += 1

        for num, question in items:
            score = assessment_answers.get(f"q{num}", "")
            if score != "":
                try:
                    score = int(score)
                except (ValueError, TypeError):
                    score = ""
            ws.cell(row=row, column=1, value=num).border = styles['thin_border']
            ws.cell(row=row, column=1).alignment = styles['center_align']
            ws.cell(row=row, column=2, value=question).border = styles['thin_border']
            ws.cell(row=row, column=2).alignment = styles['left_align']
            for i, val in enumerate([1, 2, 3, 4], start=3):
                cell = ws.cell(row=row, column=i, value="○" if score == val else "")
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align']
            ws.cell(row=row, column=7, value=score if score != "" else "").border = styles['thin_border']
            ws.cell(row=row, column=7).alignment = styles['center_align']
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="合計点").font = styles['header_font']
    ws.cell(row=row, column=2, value=data.get("total_score", ""))
    ws.cell(row=row, column=3, value="（21〜84点、31点以上で認知症の疑い）")

    row += 1
    ws.cell(row=row, column=1, value="備考").font = styles['header_font']
    ws.cell(row=row, column=2, value=data.get("remarks", ""))

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 6
    ws.column_dimensions['G'].width = 8

    save_excel_to_client_folder(wb, client_id, "dasc21")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"dasc21_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route("/api/export_dbd13", methods=["GET"])
def export_dbd13():
    """DBD-13をExcelで出力"""
    client_id = request.args.get("client_id")
    if not client_id:
        return jsonify({"status": "error", "message": "client_id が必要です"}), 400

    conn = get_connection()
    with conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM dbd13 WHERE client_id = %s ORDER BY dbd_id DESC LIMIT 1", (client_id,))
            data = cur.fetchone()

    if not data:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    import json
    assessment_answers = {}
    try:
        assessment_item = data.get("assessment_item", "")
        if assessment_item:
            assessment_answers = json.loads(assessment_item)
    except (json.JSONDecodeError, TypeError):
        assessment_answers = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "DBD-13"
    styles = create_excel_styles()

    ws.merge_cells('A1:H1')
    ws['A1'] = "DBD-13 認知症行動障害尺度（短縮版）"
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws['A1'].alignment = styles['center_align']
    ws.row_dimensions[1].height = 30

    ws['A3'] = "回答者氏名"
    ws['B3'] = data.get("respondent_name", "")
    ws['C3'] = "評価者氏名"
    ws['D3'] = data.get("evaluator_name", "")
    ws['E3'] = "記入日"
    ws['F3'] = str(data.get("entry_date", "")) if data.get("entry_date") else ""
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].border = styles['thin_border']

    ws['A5'] = "評価基準: 0=全くない 1=ほとんどない 2=ときどきある 3=よくある 4=常にある"
    ws['A5'].font = Font(italic=True, size=10)

    dbd_items = [
        ("1", "同じことを何度も何度も聞く"),
        ("2", "よく物をなくしたり、置き場所を間違えたり、隠したりしている"),
        ("3", "日常的な物事に関心を示さない"),
        ("4", "夜中に起き出す"),
        ("5", "根拠のない事を言う（例：盗まれた、配偶者が浮気している等）"),
        ("6", "昼間、寝てばかりいる"),
        ("7", "やたらに歩き回る"),
        ("8", "同じ動作をいつまでも繰り返す"),
        ("9", "口汚くののしる"),
        ("10", "場違いあるいは季節に合わない不適切な服装をする"),
        ("11", "世話をされるのを拒否する"),
        ("12", "食べられないものを口に入れる"),
        ("13", "引き出しやタンスの中身を全部出してしまう"),
    ]

    row = 7
    headers = ["No.", "評価項目", "0", "1", "2", "3", "4", "評価"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
    row += 1

    for num, question in dbd_items:
        score = assessment_answers.get(f"d{num}", "")
        if score != "":
            try:
                score = int(score)
            except (ValueError, TypeError):
                score = ""
        ws.cell(row=row, column=1, value=num).border = styles['thin_border']
        ws.cell(row=row, column=1).alignment = styles['center_align']
        ws.cell(row=row, column=2, value=question).border = styles['thin_border']
        ws.cell(row=row, column=2).alignment = styles['left_align']
        for i, val in enumerate([0, 1, 2, 3, 4], start=3):
            cell = ws.cell(row=row, column=i, value="○" if score == val else "")
            cell.border = styles['thin_border']
            cell.alignment = styles['center_align']
        ws.cell(row=row, column=8, value=score if score != "" else "").border = styles['thin_border']
        ws.cell(row=row, column=8).alignment = styles['center_align']
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="合計点").font = styles['header_font']
    ws.cell(row=row, column=2, value=data.get("total_score", ""))
    ws.cell(row=row, column=3, value="（0〜52点、点数が高いほど行動障害が重い）")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 6
    ws.column_dimensions['H'].width = 8

    save_excel_to_client_folder(wb, client_id, "dbd13")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"dbd13_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ================================
#  Flask 実行
# ================================
if __name__ == "__main__":
    app.run(debug=True)
